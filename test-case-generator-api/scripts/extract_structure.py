#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
extract_structure.py - Extract template structure from a local .xlsx file

Reads excel_template/template.xlsx using openpyxl and outputs a structure.json
containing column mapping, formatting, merged cells, and header rows data.
No Google API required — runs entirely offline.

Usage:
  python extract_structure.py
  python extract_structure.py --template excel_template/template.xlsx
  python extract_structure.py --output excel_template/structure.json
  python extract_structure.py --project-root /path/to/project

Output (excel_template/structure.json):
  {
    "templateFile": "template.xlsx",
    "sheetName": "Sheet1",
    "columnMapping": {"testSuiteName": 0, "testCaseName": 3, ...},
    "totalColumns": 13,
    "lastCol": "M",
    "headerRow": 2,
    "dataStartRow": 3,
    "columnWidths": [15, 10, ...],
    "templateRows": [...],
    "mergedCells": [...],
    "formatting": {...}
  }

Requirements:
  pip install openpyxl
"""

import argparse
import json
import os
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    print(json.dumps({"error": "Missing dependency. Run: pip install openpyxl"}))
    sys.exit(1)


# ============ PROJECT ROOT DETECTION ============
_PROJECT_MARKERS = ("catalog", "AGENTS.md", "excel_template")


def find_project_root(explicit_path=None):
    """Find project root by walking up from CWD looking for marker files."""
    if explicit_path:
        p = Path(explicit_path)
        if p.exists():
            return p

    current = Path.cwd()
    for _ in range(10):
        if any((current / m).exists() for m in _PROJECT_MARKERS):
            return current
        parent = current.parent
        if parent == current:
            break
        current = parent
    return Path.cwd()


# ============ LABEL MAPPING (from detect_template.py) ============
# Known column header labels → normalized JSON key
LABEL_TO_KEY = {
    'name':                         '__name__',
    'details':                      'details',
    'external id':                  'externalId',
    'summary':                      'summary',
    'preconditions':                'preConditions',
    'pre-conditions':               'preConditions',
    'step':                         'steps',
    'test steps':                   'testSteps',
    'expected result':              'expectedResults',
    'spec title':                   'specTitle',
    'document id':                  'documentId',
    'duration':                     'duration',
    'estimated exec. duration':     'duration',
    'estimated exec.\nduration':    'duration',
    'result':                       'result',
    'test results':                 'testResults',
    'note':                         'note',
    'notes':                        'note',
    'status':                       'status',
    'executiontype':                'executionType',
    'importance':                   'priority',
    'keywords':                     'keywords',
    'number of attachments':        'attachments',
    'actual result':                'actualResult',
    'stepexec\n type':              'stepExecType',
    'stepexectype':                 'stepExecType',
    'testcase lv1':                 'testcaseLV1',
    'testcase lv2':                 'testcaseLV2',
    'testcase lv3':                 'testcaseLV3',
    'test case id':                 'testCaseId',
    'test case title':              'testCaseTitle',
    'test case title\n(tên testcase)': 'testCaseTitle',
    'pre-conditions\n(điều kiện bắt buộc)': 'preConditions',
    'test steps\n(các bước thực hiện)': 'testSteps',
    'expected result\n(kết quả mong muốn)': 'expectedResults',
    'priority':                     'priority',
    'bugid':                        'bugId',
}

GROUP_HEADER_LABELS = {
    'test suite', 'test case', 'attachments', 'custom fields',
    'steps', 'coverage', 'time management', 'result',
    'kịch bản kiểm thử *', 'tên màn hình/tên chức năng',
}


# ============ HEADER DETECTION (from detect_template.py) ============
def score_row_as_header(row_values):
    """Score how likely a row is the column header row."""
    if not row_values:
        return 0

    score = 0
    non_empty = 0
    matched_labels = 0

    for cell_val in row_values:
        label = str(cell_val).strip().lower() if cell_val else ''
        if not label:
            continue
        non_empty += 1

        if label in GROUP_HEADER_LABELS:
            score -= 2
            continue

        if label in LABEL_TO_KEY or label == 'name':
            matched_labels += 1
            score += 3

        score += 1

    if non_empty > 0:
        score += int((matched_labels / non_empty) * 10)

    return score


def find_header_row(rows, max_rows=20):
    """Find header row index (1-based) and data start row."""
    if not rows:
        return 1, 2

    best_score = -1
    best_idx = 0

    for i, row_values in enumerate(rows[:max_rows]):
        score = score_row_as_header(row_values)
        if score > best_score:
            best_score = score
            best_idx = i

    header_row = best_idx + 1  # 1-based

    data_start = header_row + 1
    next_idx = best_idx + 1
    if next_idx < len(rows):
        next_vals = rows[next_idx]
        non_empty = [v for v in next_vals if v and str(v).strip()]
        score = score_row_as_header(next_vals)
        if 1 <= len(non_empty) <= 2 and score <= 2:
            data_start = header_row + 2

    return header_row, data_start


def build_column_mapping(row_values):
    """Build column mapping from header row values."""
    mapping = {}
    name_count = 0

    for col_idx, cell_val in enumerate(row_values):
        label = str(cell_val).strip().lower() if cell_val else ''
        if not label:
            continue

        if label == 'name':
            name_count += 1
            key = 'testSuiteName' if name_count == 1 else 'testCaseName'
            if key not in mapping:
                mapping[key] = col_idx
            continue

        key = LABEL_TO_KEY.get(label)
        if key and key != '__name__' and key not in mapping:
            mapping[key] = col_idx

    return mapping


def col_index_to_letter(index):
    """Convert 0-based column index to letter (0→A, 25→Z, 26→AA)."""
    result = ''
    index += 1
    while index > 0:
        index, remainder = divmod(index - 1, 26)
        result = chr(65 + remainder) + result
    return result


# ============ CELL PARSING ============
def argb_from_color(color):
    """Extract ARGB hex string from an openpyxl Color object."""
    if color is None:
        return None
    if color.type == 'rgb' and color.rgb and color.rgb != '00000000':
        return str(color.rgb)
    if color.type == 'indexed':
        # openpyxl indexed colors — return None for default/auto
        return None
    if color.type == 'theme':
        # Theme colors require theme resolution — skip for simplicity
        return None
    return None


def parse_cell_format(cell):
    """Extract formatting from an openpyxl cell."""
    fmt = {}

    # Background color
    fill = cell.fill
    if fill and fill.patternType and fill.patternType != 'none':
        bg = argb_from_color(fill.fgColor) or argb_from_color(fill.bgColor)
        if bg:
            fmt['backgroundColor'] = bg

    # Font
    font = cell.font
    if font:
        font_data = {}
        if font.bold:
            font_data['bold'] = True
        if font.italic:
            font_data['italic'] = True
        if font.size:
            font_data['size'] = font.size
        color = argb_from_color(font.color)
        if color:
            font_data['color'] = color
        if font.name:
            font_data['name'] = font.name
        if font_data:
            fmt['font'] = font_data

    # Alignment
    alignment = cell.alignment
    if alignment:
        align_data = {}
        if alignment.horizontal:
            align_data['horizontal'] = alignment.horizontal
        if alignment.vertical:
            align_data['vertical'] = alignment.vertical
        if alignment.wrap_text:
            align_data['wrapText'] = True
        if align_data:
            fmt['alignment'] = align_data

    # Border
    border = cell.border
    if border:
        border_data = {}
        for side_name in ('top', 'bottom', 'left', 'right'):
            side = getattr(border, side_name, None)
            if side and side.style and side.style != 'none':
                side_info = {'style': side.style}
                color = argb_from_color(side.color)
                if color:
                    side_info['color'] = color
                border_data[side_name] = side_info
        if border_data:
            fmt['border'] = border_data

    # Number format
    if cell.number_format and cell.number_format != 'General':
        fmt['numberFormat'] = cell.number_format

    return fmt


def parse_cell_value(cell):
    """Extract value from an openpyxl cell."""
    if cell.data_type == 'f':
        return None, str(cell.value) if cell.value else None  # formula
    val = cell.value
    if val is None:
        return '', None
    return str(val), None


# ============ SHEET DETECTION ============
def _scan_sheet(ws):
    """Scan a worksheet and return (row_values_list, header_row, data_start_row, column_mapping, best_score)."""
    max_scan = min(20, ws.max_row or 20)
    row_values_list = []
    for row_idx in range(1, max_scan + 1):
        vals = []
        for col_idx in range(1, (ws.max_column or 1) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            vals.append(cell.value)
        row_values_list.append(vals)

    header_row, data_start_row = find_header_row(row_values_list)
    header_vals = row_values_list[header_row - 1] if header_row <= len(row_values_list) else []
    column_mapping = build_column_mapping(header_vals)

    # Score = number of mapped columns (more mapped columns = more likely the test case sheet)
    best_score = len(column_mapping)
    return row_values_list, header_row, data_start_row, column_mapping, best_score


def find_best_sheet(wb, sheet_name_hint=None):
    """Find the best worksheet for test case data.

    Priority:
    1. Explicit sheet name (--sheet)
    2. Sheet with the most recognized column headers
    3. First sheet (fallback)
    """
    if sheet_name_hint:
        if sheet_name_hint in wb.sheetnames:
            return wb[sheet_name_hint]
        # Try case-insensitive match
        for name in wb.sheetnames:
            if name.lower() == sheet_name_hint.lower():
                return wb[name]
        # Not found — fall through to auto-detect

    best_ws = wb.worksheets[0]
    best_score = -1

    for ws in wb.worksheets:
        _, _, _, mapping, score = _scan_sheet(ws)
        if score > best_score:
            best_score = score
            best_ws = ws

    return best_ws


# ============ MAIN EXTRACTION ============
def extract_structure(template_path, output_path=None, sheet_name_hint=None):
    """Extract full template structure from .xlsx file."""
    if not os.path.isfile(template_path):
        return {"error": f"Template file not found: {template_path}"}

    wb = openpyxl.load_workbook(template_path, data_only=False)
    ws = find_best_sheet(wb, sheet_name_hint)
    sheet_name = ws.title

    # Read all rows as values for header detection (max 20 rows)
    row_values_list, header_row, data_start_row, column_mapping, _ = _scan_sheet(ws)
    header_vals = row_values_list[header_row - 1] if header_row <= len(row_values_list) else []

    # Total columns: count up to last non-empty cell in header row
    total_columns = len(header_vals)
    while total_columns > 0 and not (header_vals[total_columns - 1] and str(header_vals[total_columns - 1]).strip()):
        total_columns -= 1
    if total_columns == 0:
        total_columns = len(header_vals)

    last_col = col_index_to_letter(total_columns - 1)

    # Extract column widths
    column_widths = []
    for col_idx in range(1, total_columns + 1):
        col_letter = get_column_letter(col_idx)
        dim = ws.column_dimensions.get(col_letter)
        width = dim.width if dim and dim.width else 15
        column_widths.append(round(width, 2))

    # Parse template header rows (1 to dataStartRow - 1) with full formatting
    template_end_row = data_start_row - 1
    template_rows = []
    for row_idx in range(1, template_end_row + 1):
        row_obj = ws.row_dimensions.get(row_idx)
        row_height = row_obj.height if row_obj and row_obj.height else None

        cells_data = []
        for col_idx in range(1, total_columns + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            value, formula = parse_cell_value(cell)
            cell_fmt = parse_cell_format(cell)

            cell_data = {"value": value or ""}
            if formula:
                cell_data["formula"] = formula
            if cell_fmt:
                cell_data.update(cell_fmt)

            cells_data.append(cell_data)

        row_data = {"rowNumber": row_idx, "cells": cells_data}
        if row_height:
            row_data["height"] = row_height
        template_rows.append(row_data)

    # Detect merged cells in header area
    merged_cells = []
    for mc in ws.merged_cells.ranges:
        # Only include merges that overlap with the header area
        if mc.min_row <= template_end_row:
            merged_cells.append({
                "startRow": mc.min_row,
                "startCol": mc.min_col,
                "endRow": min(mc.max_row, template_end_row),
                "endCol": mc.max_col,
            })

    # Default formatting for data rows
    formatting = {
        "suiteHeader": {
            "backgroundColor": "FFDAEAD0",
            "font": {"bold": True, "color": "FF000000", "size": 11},
            "merge": True,
        },
        "testCaseRow": {
            "backgroundColor": "FFFFFFFF",
            "font": {"bold": False, "color": "FF000000", "size": 11},
            "wrapText": True,
        },
    }

    result = {
        "templateFile": os.path.basename(template_path),
        "sheetName": sheet_name,
        "columnMapping": column_mapping,
        "totalColumns": total_columns,
        "lastCol": last_col,
        "headerRow": header_row,
        "dataStartRow": data_start_row,
        "columnWidths": column_widths,
        "templateRows": template_rows,
        "mergedCells": merged_cells,
        "formatting": formatting,
    }

    # Write output
    if output_path:
        os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(result, f, ensure_ascii=False, indent=2)
        print(json.dumps({
            "success": True,
            "outputFile": str(output_path),
            "sheetName": sheet_name,
            "columnMapping": column_mapping,
            "totalColumns": total_columns,
            "headerRow": header_row,
            "dataStartRow": data_start_row,
        }, ensure_ascii=False))
    else:
        print(json.dumps(result, ensure_ascii=False, indent=2))

    return result


def main():
    parser = argparse.ArgumentParser(
        description='Extract template structure from a local .xlsx file to structure.json'
    )
    parser.add_argument('--template', help='Path to .xlsx template file (default: excel_template/template.xlsx)')
    parser.add_argument('--output', help='Output path for structure.json (default: excel_template/structure.json)')
    parser.add_argument('--sheet', help='Sheet name to extract (auto-detected if omitted)')
    parser.add_argument('--project-root', help='Explicit project root path (auto-detected if omitted)')
    args = parser.parse_args()

    project_root = find_project_root(args.project_root)

    template_path = args.template or str(project_root / 'excel_template' / 'template.xlsx')
    output_path = args.output or str(project_root / 'excel_template' / 'structure.json')

    try:
        extract_structure(template_path, output_path, sheet_name_hint=args.sheet)
    except Exception as e:
        print(json.dumps({"error": str(e)}))
        sys.exit(1)


if __name__ == '__main__':
    main()
