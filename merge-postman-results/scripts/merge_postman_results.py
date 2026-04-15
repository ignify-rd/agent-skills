#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
merge_postman_results.py — Merge auto-postman API test results into a new Google Spreadsheet.

Reads api_results_*.xlsx (auto-postman output), downloads the target Google Sheet as xlsx
(used as a template — it is NOT modified), fills Actual Result with the response JSON body,
evaluates Pass/Fail by comparing status codes and writes the verdict to the "Kết quả hiện tại"
(status) column, embeds screenshots in an Evidence sheet, and creates a BRAND-NEW Google Sheet
with the merged data.

Usage:
  python merge_postman_results.py \
    --source api_results_20260415_091720.xlsx \
    --target "https://docs.google.com/spreadsheets/d/SPREADSHEET_ID/edit" \
    --structure structure.json \
    [--name "My Merged Results 2026-04-15"] \
    [--dry-run]

Requirements:
  pip install openpyxl pillow google-api-python-client google-auth google-auth-oauthlib
"""

import argparse
import json
import os
import re
import sys
import tempfile
import time
from pathlib import Path

# Force UTF-8 output on Windows
if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if sys.stderr.encoding and sys.stderr.encoding.lower() != "utf-8":
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")


# ---------------------------------------------------------------------------
# Google OAuth2 credentials
# ---------------------------------------------------------------------------

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]

CREDS_PATH = Path.home() / "google-sheets-mcp" / "dist" / ".gsheets-server-credentials.json"
OAUTH_PATH = Path.home() / "google-sheets-mcp" / "dist" / "gcp-oauth.keys.json"


def _save_credentials(creds):
    """Save OAuth2 credentials back to MCP server's credential file."""
    data = {
        "access_token": creds.token,
        "refresh_token": creds.refresh_token,
        "scope": " ".join(creds.scopes or SCOPES),
        "token_type": "Bearer",
        "expiry_date": (
            int(creds.expiry.timestamp() * 1000) if creds.expiry
            else int(time.time() * 1000) + 3600000
        ),
    }
    with open(CREDS_PATH, "w") as f:
        json.dump(data, f, indent=2)
    print(f"  [auth] Credentials saved to {CREDS_PATH}")


def get_google_credentials(creds_path=None, oauth_path=None):
    """
    Load OAuth2 credentials with Sheets + Drive scope.

    Reuses the MCP gsheets server's stored credentials. If the current token
    doesn't have Drive scope, triggers a one-time browser re-authorization.
    """
    from google.oauth2.credentials import Credentials
    from google.auth.transport.requests import Request

    cp = Path(creds_path) if creds_path else CREDS_PATH
    op = Path(oauth_path) if oauth_path else OAUTH_PATH

    if not op.exists():
        raise FileNotFoundError(
            f"OAuth client keys not found: {op}\n"
            "Set up the gsheets MCP server first, or pass --oauth-keys."
        )

    with open(op) as f:
        oauth_keys = json.load(f)["installed"]

    creds = None

    # Try loading stored credentials
    if cp.exists():
        with open(cp) as f:
            stored = json.load(f)

        creds = Credentials(
            token=stored.get("access_token"),
            refresh_token=stored.get("refresh_token"),
            token_uri=oauth_keys["token_uri"],
            client_id=oauth_keys["client_id"],
            client_secret=oauth_keys["client_secret"],
            scopes=SCOPES,
        )

        if creds.expired or not creds.valid:
            try:
                creds.refresh(Request())
                _save_credentials(creds)
            except Exception:
                creds = None  # will re-auth below

    # Verify Drive scope works
    if creds and creds.valid:
        try:
            from googleapiclient.discovery import build
            drive = build("drive", "v3", credentials=creds, cache_discovery=False)
            drive.about().get(fields="user").execute()
            return creds
        except Exception:
            print("  [auth] Current token lacks Drive scope — re-authorization needed.")
            creds = None

    # Re-authorize with full scopes (opens browser)
    print("[auth] Opening browser for Google OAuth (Sheets + Drive) ...")
    from google_auth_oauthlib.flow import InstalledAppFlow
    flow = InstalledAppFlow.from_client_secrets_file(str(op), SCOPES)
    creds = flow.run_local_server(port=0)
    _save_credentials(creds)
    return creds


def parse_spreadsheet_id(url_or_id: str) -> str:
    """Extract spreadsheet ID from a Google Sheets URL or return as-is."""
    m = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", url_or_id)
    return m.group(1) if m else url_or_id


# ---------------------------------------------------------------------------
# Google Drive download / upload
# ---------------------------------------------------------------------------

def download_gsheet_as_xlsx(spreadsheet_id: str, creds, output_path: str):
    """Download a Google Sheet as .xlsx via authenticated HTTP export."""
    import urllib.request

    export_url = (
        f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=xlsx"
    )
    req = urllib.request.Request(export_url)
    req.add_header("Authorization", f"Bearer {creds.token}")

    with urllib.request.urlopen(req) as resp, open(output_path, "wb") as f:
        f.write(resp.read())

    size_kb = os.path.getsize(output_path) / 1024
    print(f"  Downloaded → {output_path} ({size_kb:.0f} KB)")


def upload_xlsx_to_gsheet(xlsx_path: str, spreadsheet_id: str, creds) -> str:
    """
    Upload merged .xlsx back to the Google Sheet, replacing its content.

    Uses Drive API files.update with xlsx media body. The body mimeType parameter
    instructs Drive to convert the uploaded xlsx to native Google Sheets format.
    """
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaFileUpload

    drive = build("drive", "v3", credentials=creds, cache_discovery=False)

    media = MediaFileUpload(
        xlsx_path,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        resumable=False,
    )

    drive.files().update(
        fileId=spreadsheet_id,
        media_body=media,
        body={"mimeType": "application/vnd.google-apps.spreadsheet"},
    ).execute()

    url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/edit"
    print(f"  Uploaded → {url}")
    return url


def create_new_gsheet_from_xlsx(xlsx_path: str, title: str, creds) -> str:
    """
    Create a brand-new Google Sheet from the merged .xlsx file.

    Uses Drive API files.create to make a fresh spreadsheet, then uploads the xlsx
    into it (converted to native Google Sheets format). Returns the new spreadsheet URL.
    """
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaFileUpload

    drive = build("drive", "v3", credentials=creds, cache_discovery=False)

    # 1. Create empty spreadsheet with the given title
    file_metadata = {
        "name": title,
        "mimeType": "application/vnd.google-apps.spreadsheet",
    }
    new_file = drive.files().create(body=file_metadata, fields="id").execute()
    new_spreadsheet_id = new_file.get("id")
    print(f"  Created new spreadsheet → ID: {new_spreadsheet_id}")

    # 2. Upload xlsx into the new spreadsheet (Drive converts to native format)
    media = MediaFileUpload(
        xlsx_path,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        resumable=False,
    )
    drive.files().update(
        fileId=new_spreadsheet_id,
        media_body=media,
        body={"mimeType": "application/vnd.google-apps.spreadsheet"},
    ).execute()

    url = f"https://docs.google.com/spreadsheets/d/{new_spreadsheet_id}/edit"
    print(f"  Uploaded → {url}")
    return url, new_spreadsheet_id


# ---------------------------------------------------------------------------
# Extract embedded images from xlsx drawing layer
# ---------------------------------------------------------------------------

def _extract_embedded_images(xlsx_path: str, col_index: int = None) -> dict:
    """
    Parse the drawing XML inside an xlsx (zip) file and return a dict mapping
    0-based row index → raw image bytes.

    Drawing coordinates use 0-based row/col indices where Excel row N = drawing row N-1.

    Args:
        xlsx_path:  Path to the xlsx file.
        col_index:  0-based column index to filter images by (None = all columns).

    Returns:
        dict: { drawing_row_0based: bytes }
    """
    import zipfile
    import xml.etree.ElementTree as ET

    NS_SD = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
    NS_A  = "http://schemas.openxmlformats.org/drawingml/2006/main"
    NS_R  = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    NS_PKG = "http://schemas.openxmlformats.org/package/2006/relationships"

    result = {}

    with zipfile.ZipFile(xlsx_path) as zf:
        names = set(zf.namelist())

        drawing_files = sorted(
            n for n in names
            if n.startswith("xl/drawings/drawing") and n.endswith(".xml") and "_rels" not in n
        )
        if not drawing_files:
            return result

        for drawing_path in drawing_files:
            rels_path = f"xl/drawings/_rels/{drawing_path.split('/')[-1]}.rels"
            if rels_path not in names:
                continue

            rels_root = ET.fromstring(zf.read(rels_path).decode("utf-8"))
            rid_to_media = {}
            for rel in rels_root.findall(f"{{{NS_PKG}}}Relationship"):
                rid    = rel.get("Id")
                target = rel.get("Target", "")
                if "media" in target:
                    media_path = "xl/media/" + target.rsplit("/", 1)[-1]
                    rid_to_media[rid] = media_path

            drawing_root = ET.fromstring(zf.read(drawing_path).decode("utf-8"))
            anchor_types = [f"{{{NS_SD}}}oneCellAnchor", f"{{{NS_SD}}}twoCellAnchor"]

            for anchor_tag in anchor_types:
                for anchor in drawing_root.findall(anchor_tag):
                    from_el = anchor.find(f"{{{NS_SD}}}from")
                    if from_el is None:
                        continue

                    col_el = from_el.find(f"{{{NS_SD}}}col")
                    row_el = from_el.find(f"{{{NS_SD}}}row")
                    if col_el is None or row_el is None:
                        continue

                    img_col = int(col_el.text)
                    img_row = int(row_el.text)

                    if col_index is not None and img_col != col_index:
                        continue

                    pic = anchor.find(f"{{{NS_SD}}}pic")
                    if pic is None:
                        continue
                    blip_fill = pic.find(f"{{{NS_SD}}}blipFill")
                    if blip_fill is None:
                        continue
                    blip = blip_fill.find(f"{{{NS_A}}}blip")
                    if blip is None:
                        continue

                    rid = blip.get(f"{{{NS_R}}}embed")
                    if not rid:
                        continue

                    media_path = rid_to_media.get(rid)
                    if media_path and media_path in names:
                        result[img_row] = zf.read(media_path)

    return result


# ---------------------------------------------------------------------------
# Load source (auto-postman output)
# ---------------------------------------------------------------------------

def load_source(source_path: str) -> dict:
    """
    Load auto-postman results xlsx.

    Supports two screenshot formats:
      - NEW: screenshots embedded directly as images in the xlsx drawing layer.
      - OLD: Screenshot cell contains a relative file path to a .png file.

    Returns:
        dict: { api_name: { status_code, response_body, screenshot } }
              'screenshot' is raw bytes (embedded) or a file-path string,
              or None when no screenshot is available.
    """
    import openpyxl

    wb = openpyxl.load_workbook(source_path, data_only=True)
    ws = wb.active
    source_dir = Path(source_path).parent

    header_row_idx = None
    col_indices = {}

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        vals = [str(v).strip() if v is not None else "" for v in row]
        if "API Name" in vals:
            header_row_idx = row_idx
            for col_idx, val in enumerate(vals):
                col_indices[val] = col_idx
            break

    if header_row_idx is None:
        raise ValueError(
            f"Cannot find 'API Name' header in source xlsx: {source_path}"
        )

    api_col    = col_indices.get("API Name")
    status_col = col_indices.get("Status Code")
    body_col   = col_indices.get("Response Body")
    shot_col   = col_indices.get("Screenshot")

    embedded = {}
    if shot_col is not None:
        embedded = _extract_embedded_images(source_path, col_index=shot_col)

    results = {}

    for excel_row, row in enumerate(
        ws.iter_rows(min_row=header_row_idx + 1, values_only=True),
        start=header_row_idx + 1,
    ):
        if api_col is None or row[api_col] is None:
            continue

        name = str(row[api_col]).strip()
        if not name:
            continue

        drawing_row = excel_row - 1
        screenshot = embedded.get(drawing_row)

        if screenshot is None and shot_col is not None and row[shot_col]:
            raw = str(row[shot_col]).replace("\\", "/").strip()
            for candidate in [
                source_dir / raw,
                source_dir.parent / raw,
                Path(raw),
            ]:
                if candidate.exists():
                    screenshot = str(candidate)
                    break

        results[name] = {
            "status_code": (
                str(row[status_col]).strip()
                if status_col is not None and row[status_col] is not None
                else ""
            ),
            "response_body": (
                str(row[body_col]).replace("\xa0", " ").strip()
                if body_col is not None and row[body_col] is not None
                else ""
            ),
            "screenshot": screenshot,
        }

    return results


# ---------------------------------------------------------------------------
# Column detection from structure.json + header row scan
# ---------------------------------------------------------------------------

FIELD_SYNONYMS = {
    "testCaseName": [
        "Name", "Test Case Name", "TestCase Name", "Tên test case",
        "API Name", "testCaseName", "Name Testcase ",
    ],
    "status": [
        "Status", "Trạng thái", "Kết quả", "Result",
        "Pass/Fail", "ExecutionStatus", "Kết quả hiện tại",
    ],
    "expectedResults": [
        "Expected Result", "Expected Results", "Expected",
        "Kết quả mong đợi", "ExpectedResult",
    ],
    "actualResult": [
        "Actual Result", "Actual", "Kết quả thực tế", "ActualResult",
    ],
}


def detect_columns(ws, structure: dict) -> dict:
    """
    Merge structure.json columnMapping with header-row auto-detection.

    Returns:
        dict: { field_name: 0-based column index }
    """
    mapping = {k: int(v) for k, v in structure.get("columnMapping", {}).items()}
    header_row = structure.get("headerRow", 1)

    headers = [cell.value for cell in ws[header_row]]

    # Also scan the row above the header for merged-cell labels
    row_above = []
    if header_row > 1:
        row_above = [cell.value for cell in ws[header_row - 1]]

    for field, synonyms in FIELD_SYNONYMS.items():
        if field not in mapping:
            for col_idx, header_val in enumerate(headers):
                if header_val in synonyms:
                    mapping[field] = col_idx
                    break
            # If still not found, check the merged-cell row above
            if field not in mapping:
                for col_idx, header_val in enumerate(row_above):
                    if header_val in synonyms:
                        mapping[field] = col_idx
                        break

    return mapping


def auto_detect_structure(ws, structure: dict) -> tuple:
    """
    Scan the entire worksheet to find the real header row and column mapping
    when the structure.json values appear to be wrong (e.g., matched = 0).

    Returns:
        (header_row, data_start_row, col_map)
    """
    all_synonyms = {v: field for field, syns in FIELD_SYNONYMS.items() for v in syns}

    best_row     = None
    best_score   = 0
    best_mapping = {}

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        found = {}
        for col_idx, val in enumerate(row):
            if val and str(val).strip() in all_synonyms:
                field = all_synonyms[str(val).strip()]
                if field not in found:
                    found[field] = col_idx

        if len(found) >= 2 and len(found) > best_score:
            best_score   = len(found)
            best_row     = row_idx
            best_mapping = found

    if best_row is None:
        orig_header = structure.get("headerRow", 1)
        return orig_header, structure.get("dataStartRow", orig_header + 1), detect_columns(ws, structure)

    existing = {k: int(v) for k, v in structure.get("columnMapping", {}).items()}
    merged = {**existing, **best_mapping}

    data_row = best_row + 1
    tc_col = merged.get("testCaseName")
    if tc_col is not None:
        tc_val = ws.cell(data_row, tc_col + 1).value
        if not tc_val:
            data_row = best_row + 2

    return best_row, data_row, merged


# ---------------------------------------------------------------------------
# JSON replacement in expected result text
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# Cell formatting helpers
# ---------------------------------------------------------------------------

def _wrap_align():
    """Return an Alignment with wrap_text=True, top-aligned."""
    from openpyxl.styles import Alignment
    return Alignment(wrap_text=True, vertical="top")


def _black_font():
    """Return a Font with black color."""
    from openpyxl.styles import Font
    return Font(color="000000")


def _white_fill():
    """Return a PatternFill with white background."""
    from openpyxl.styles import PatternFill
    return PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")


def _autofit_row_height(ws, row_idx: int, *col_indices):
    """Set row height based on the maximum line count across the given columns."""
    LINE_HEIGHT_PT = 15
    max_lines = 1
    for col_idx in col_indices:
        if col_idx is None:
            continue
        val = ws.cell(row=row_idx, column=col_idx + 1).value
        if val:
            lines = str(val).count("\n") + 1
            max_lines = max(max_lines, lines)
    height = min(max_lines * LINE_HEIGHT_PT, 409)
    ws.row_dimensions[row_idx].height = height


# ---------------------------------------------------------------------------
# Pass / Fail evaluation
# ---------------------------------------------------------------------------

def _extract_status_code(text: str):
    """Pull the first 3-digit status code from a text block."""
    m = re.search(r"\b(\d{3})\b", text or "")
    return m.group(1) if m else None


def evaluate_simple(expected_text: str, actual_text: str):
    """
    Rule-based evaluation: compare status codes extracted from both texts.

    Returns:
        (verdict, reason)  — verdict is 'PASS', 'FAIL', or 'N/A'
    """
    if not actual_text or not actual_text.strip():
        return "N/A", "No actual result"

    exp_status = _extract_status_code(expected_text)
    act_status = _extract_status_code(actual_text)

    if exp_status and act_status:
        if exp_status == act_status:
            return "PASS", f"Status code match: {act_status}"
        return "FAIL", f"Status mismatch: expected {exp_status}, got {act_status}"

    return "FAIL", "Cannot determine (no 3-digit status code found)"



# ---------------------------------------------------------------------------
# Evidence sheet (embedded images in xlsx)
# ---------------------------------------------------------------------------

def update_evidence_sheet(wb, source_results: dict):
    """
    Create (or replace) the 'Evidence' sheet with embedded screenshot images.

    'screenshot' in source_results can be:
        bytes      — image extracted from the source xlsx drawing layer
        str        — absolute file path
        None       — no screenshot available
    """
    SHEET_NAME = "Evidence"

    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]

    ws = wb.create_sheet(SHEET_NAME, index=1)
    ws["A1"] = "ID"
    ws["B1"] = "Evidence"
    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["B"].width = 90

    row_num = 2
    for api_name, data in source_results.items():
        # Only include matched test cases (externalId is set during merge)
        if "externalId" not in data:
            continue
        display_id = data["externalId"]
        ws[f"A{row_num}"] = display_id
        shot = data.get("screenshot")

        if isinstance(shot, (bytes, bytearray)) and shot:
            ok = _embed_image_bytes(ws, row_num, "B", shot)
            if not ok:
                ws[f"B{row_num}"] = "[Image embed failed]"
        elif isinstance(shot, str) and os.path.exists(shot):
            _embed_image_file(ws, row_num, "B", shot)
        else:
            ws[f"B{row_num}"] = "[No screenshot]"

        row_num += 1


def _scale_dimensions(width: int, height: int, max_w: int = 700, max_h: int = 450):
    """Return (w, h) scaled down to fit within max_w x max_h, preserving aspect ratio."""
    scale = min(max_w / max(width, 1), max_h / max(height, 1), 1.0)
    return int(width * scale), int(height * scale)


def _apply_image_to_cell(ws, img, row_num: int, col_letter: str):
    """Anchor img to the cell, set row height and column width to match."""
    from openpyxl.utils import column_index_from_string, get_column_letter

    img.anchor = f"{col_letter}{row_num}"
    ws.add_image(img)

    ws.row_dimensions[row_num].height = img.height / 1.33
    col_idx = column_index_from_string(col_letter)
    ws.column_dimensions[get_column_letter(col_idx)].width = max(
        ws.column_dimensions[get_column_letter(col_idx)].width or 0,
        img.width / 7,
    )


def _embed_image_bytes(ws, row_num: int, col_letter: str, img_bytes: bytes) -> bool:
    """Embed raw image bytes into a cell. Returns True on success."""
    try:
        import io
        from openpyxl.drawing.image import Image as XLImage
        from PIL import Image as PILImage

        pil_img = PILImage.open(io.BytesIO(img_bytes))
        w, h = _scale_dimensions(pil_img.width, pil_img.height)

        buf = io.BytesIO()
        fmt = pil_img.format or "PNG"
        pil_img.save(buf, format=fmt)
        buf.seek(0)

        xl_img = XLImage(buf)
        xl_img.width, xl_img.height = w, h
        _apply_image_to_cell(ws, xl_img, row_num, col_letter)
        return True

    except Exception as exc:
        ws[f"{col_letter}{row_num}"] = f"[Image error: {exc}]"
        return False


def _embed_image_file(ws, row_num: int, col_letter: str, img_path: str):
    """Embed an image from a file path into a cell."""
    try:
        from openpyxl.drawing.image import Image as XLImage

        xl_img = XLImage(img_path)
        w, h = _scale_dimensions(xl_img.width, xl_img.height)
        xl_img.width, xl_img.height = w, h
        _apply_image_to_cell(ws, xl_img, row_num, col_letter)

    except Exception as exc:
        ws[f"{col_letter}{row_num}"] = f"[Image error: {exc}]"


# ---------------------------------------------------------------------------
# Verification
# ---------------------------------------------------------------------------

def _verify_output(out_path: str, structure: dict, col_map: dict, data_start_row: int, expected_matched: int) -> dict:
    """
    Re-open the saved file and verify key columns were written correctly.

    Checks:
      - actualResult: how many rows have response content written
      - status: how many rows have PASS/FAIL/N/A verdict written
      - expectedResults: how many rows where JSON was replaced (contains '{')
      - Evidence sheet: row count vs image count
    """
    import openpyxl

    wb = openpyxl.load_workbook(out_path, data_only=True)
    ws = wb.worksheets[0]

    act_col = col_map.get("actualResult")
    exp_col = col_map.get("expectedResults")

    actual_written = 0
    status_written = 0
    verdict_counts = {"PASS": 0, "FAIL": 0, "N/A": 0}
    json_replaced   = 0
    blank_actual    = []

    verdicts_set = {"PASS", "FAIL", "N/A"}
    sts_col_v = col_map.get("status")

    for row_idx in range(data_start_row, ws.max_row + 1):
        tc_col = col_map.get("testCaseName")
        if tc_col is not None:
            tc_val = ws.cell(row=row_idx, column=tc_col + 1).value
            if not tc_val:
                continue

        if act_col is not None:
            cell_val = ws.cell(row=row_idx, column=act_col + 1).value
            text = str(cell_val or "").strip()
            if text:
                actual_written += 1
            else:
                blank_actual.append(row_idx)

        # Check status column for verdict
        if sts_col_v is not None:
            sts_val = ws.cell(row=row_idx, column=sts_col_v + 1).value
            sts_text = str(sts_val or "").strip()
            if sts_text in verdicts_set:
                status_written += 1
                verdict_counts[sts_text] = verdict_counts.get(sts_text, 0) + 1

        if exp_col is not None:
            exp_val = ws.cell(row=row_idx, column=exp_col + 1).value
            if exp_val and "{" in str(exp_val):
                json_replaced += 1

    ev_rows   = 0
    ev_images = 0
    if "Evidence" in wb.sheetnames:
        ev_ws = wb["Evidence"]
        ev_rows   = max(ev_ws.max_row - 1, 0)
        ev_images = len(ev_ws._images)

    ok = actual_written == expected_matched

    return {
        "ok":              ok,
        "expected_matched": expected_matched,
        "actual_written":  actual_written,
        "status_written":  status_written,
        "verdict_counts":  verdict_counts,
        "json_replaced":   json_replaced,
        "evidence_rows":   ev_rows,
        "evidence_images": ev_images,
        "blank_actual_rows": blank_actual,
    }


def _print_verify(v: dict):
    ok_mark = "✓" if v["ok"] else "✗"
    print(f"  {ok_mark} actualResult written : {v['actual_written']} / {v['expected_matched']}"
          + ("" if v["ok"] else "  ← MISMATCH"))
    print(f"  ✓ status (Kết quả hiện tại) written : {v.get('status_written', 0)}")
    print(f"    PASS={v['verdict_counts'].get('PASS',0)}  "
          f"FAIL={v['verdict_counts'].get('FAIL',0)}  "
          f"N/A={v['verdict_counts'].get('N/A',0)}")
    print(f"  ✓ expectedResults with JSON : {v['json_replaced']} rows")
    ev_ok = v["evidence_images"] == v["evidence_rows"]
    ev_mark = "✓" if ev_ok else "✗"
    print(f"  {ev_mark} Evidence sheet : {v['evidence_images']} images / {v['evidence_rows']} rows"
          + ("" if ev_ok else "  ← some images missing"))
    if v["blank_actual_rows"]:
        print(f"  [warn] Rows with blank actual result: {v['blank_actual_rows']}")
    if not v["ok"] or not ev_ok:
        print("  → Check structure.json column mapping and re-run.")


# ---------------------------------------------------------------------------
# ExternalId-based fallback merge
# ---------------------------------------------------------------------------

def _run_externalid_merge_loop(ws, ws_ro, source_results, col_map, data_start_row, dry_run):
    """
    Fallback merge when name-based matching yields zero results.

    Target has an 'externalId' column (e.g. "API1", "API2", ...) that matches
    source API Name values exactly.  This function:
      1. Collects target rows that have an externalId value (skip section headers).
      2. Collects source results that are NOT "Login" or "New Request".
      3. Zips them in order (positional match).
      4. Writes verdict and externalId into source_results so update_evidence_sheet
         can filter by it.
    """
    rd = ws_ro or ws

    tc_col  = col_map.get("testCaseName")
    exp_col = col_map.get("expectedResults")
    act_col = col_map.get("actualResult")
    sts_col = col_map.get("status")
    ext_col = col_map.get("externalId")

    verdicts = {"PASS": 0, "FAIL": 0, "N/A": 0}
    matched  = 0

    # Collect target rows that have an externalId (these are real test case rows)
    target_rows = []
    for row_idx in range(data_start_row, ws.max_row + 1):
        ext_val = rd.cell(row=row_idx, column=ext_col + 1).value if ext_col is not None else None
        if ext_val is None:
            continue
        ext_id = str(ext_val).strip()
        if not ext_id:
            continue
        target_rows.append(row_idx)

    # Collect source entries excluding Login and New Request
    SKIP_NAMES = {"login", "new request"}
    source_list = [
        (name, data)
        for name, data in source_results.items()
        if name.lower().strip() not in SKIP_NAMES
    ]

    count = min(len(target_rows), len(source_list))
    print(f"  [externalId fallback] {len(target_rows)} target rows × {len(source_list)} "
          f"source entries → {count} positional matches")

    for i in range(count):
        target_row, (api_name, data) = target_rows[i], source_list[i]

        matched += 1
        data["externalId"] = api_name

        expected_text = ""
        if exp_col is not None:
            exp_val = rd.cell(row=target_row, column=exp_col + 1).value
            expected_text = str(exp_val or "")

        raw_actual = data["response_body"]
        verdict, reason = evaluate_simple(expected_text, raw_actual)
        verdicts[verdict] = verdicts.get(verdict, 0) + 1

        tc_val = rd.cell(row=target_row, column=tc_col + 1).value
        tc_display = str(tc_val or api_name)[:65]
        print(f"  [{verdict:7s}] {tc_display}")

        if dry_run:
            continue

        if act_col is not None:
            cell = ws.cell(row=target_row, column=act_col + 1)
            cell.value = raw_actual
            cell.alignment = _wrap_align()
            cell.font = _black_font()
            cell.fill = _white_fill()

        if sts_col is not None:
            cell = ws.cell(row=target_row, column=sts_col + 1)
            cell.value = verdict
            cell.alignment = _wrap_align()
            cell.font = _black_font()
            cell.fill = _white_fill()

        _autofit_row_height(ws, target_row, act_col)

    print()
    print(f"  [externalId fallback] Matched : {matched} / {len(source_results)}")
    print(f"  PASS    : {verdicts.get('PASS', 0)}")
    print(f"  FAIL    : {verdicts.get('FAIL', 0)}")
    print(f"  N/A     : {verdicts.get('N/A', 0)}")

    return matched, [], verdicts


# ---------------------------------------------------------------------------
# Main merge logic
# ---------------------------------------------------------------------------

def _run_merge_loop(ws, source_results, col_map, data_start_row, dry_run, ws_readonly=None):
    """
    Core merge loop — matches source results to target rows and writes data.

    Matches target testCaseName against source_results keys.
    On every successful match, stores `externalId` back into source_results[api_name]
    so that update_evidence_sheet() can filter by it.

    Args:
        ws:           Writable worksheet (may contain formulas as strings).
        ws_readonly:  data_only worksheet for reading computed values (optional).
                      When provided, cell values are read from ws_readonly but
                      writes go to ws.  This handles formula columns correctly.

    Writes:
      - actualResult column: response JSON body from source
      - status column ("Kết quả hiện tại"): PASS / FAIL / N/A verdict

    Returns:
        (matched, not_found, verdicts)
    """
    rd = ws_readonly or ws  # reader worksheet

    tc_col  = col_map.get("testCaseName")
    exp_col = col_map.get("expectedResults")
    act_col = col_map.get("actualResult")
    sts_col = col_map.get("status")
    ext_col = col_map.get("externalId")

    matched = 0
    not_found = []
    verdicts = {"PASS": 0, "FAIL": 0, "N/A": 0}
    row_verdicts = {}  # {row_1based: verdict} for apply_cell_background_colors

    # Build reverse lookup: target ID value → source_results key
    # We scan every column to find one whose values match source API names.
    # The correct ID column may be the first column (A) or any other column
    # depending on the sheet layout.
    target_id_col = None
    for scan_col_idx in range(0, ws.max_column):
        vals = [rd.cell(row=row_idx, column=scan_col_idx + 1).value for row_idx in range(data_start_row, min(data_start_row + 10, ws.max_row + 1))]
        matched_count = sum(1 for v in vals if v and str(v).strip() in source_results)
        if matched_count >= 2:
            target_id_col = scan_col_idx
            print(f"      [match-hint] ID column found at col {scan_col_idx} (matched {matched_count} source keys)")
            break

    if target_id_col is None and ext_col is not None:
        # Fall back to the configured externalId column
        target_id_col = ext_col

    # Rebuild ext_to_source_key with whichever column has more matches
    def _build_lookup(col_idx):
        lut = {}
        for row_idx in range(data_start_row, ws.max_row + 1):
            ext_val = rd.cell(row=row_idx, column=col_idx + 1).value
            if ext_val:
                key = str(ext_val).strip()
                if key in source_results and key not in lut:
                    lut[key] = key
        return lut

    if target_id_col is not None and ext_col is not None and ext_col != target_id_col:
        # structure.json externalId col may be wrong for this sheet
        orig_lut = _build_lookup(ext_col)
        new_lut  = _build_lookup(target_id_col)
        if len(orig_lut) >= len(new_lut):
            print(f"      [match-hint] structure.json externalId col={ext_col} "
                  f"({len(orig_lut)} matches) >= scanned col={target_id_col} "
                  f"({len(new_lut)} matches) — using structure.json")
            target_id_col = ext_col
            ext_to_source_key = orig_lut
        else:
            print(f"      [match-hint] structure.json externalId col={ext_col} "
                  f"({len(orig_lut)} matches) < scanned col={target_id_col} "
                  f"({len(new_lut)} matches) — using scanned column")
            ext_to_source_key = new_lut
    else:
        ext_to_source_key = _build_lookup(target_id_col or 0)

    for row_idx in range(data_start_row, ws.max_row + 1):
        tc_val = rd.cell(row=row_idx, column=tc_col + 1).value
        if tc_val is None:
            continue
        tc_name = str(tc_val).strip()
        if not tc_name:
            continue

        source_key = tc_name

        # Primary match: tc_name matches source_results key directly
        if source_key not in source_results:
            # Fallback: try matching via scanned ID column (e.g. col A = "API1")
            source_key = None
            if target_id_col is not None:
                id_val = rd.cell(row=row_idx, column=target_id_col + 1).value
                if id_val:
                    source_key = ext_to_source_key.get(str(id_val).strip())
            if source_key is None:
                not_found.append(tc_name)
                continue

        matched += 1
        data = source_results[source_key]

        # Store externalId for update_evidence_sheet filter
        id_write_col = target_id_col if target_id_col is not None else ext_col
        if id_write_col is not None:
            ext_id_val = rd.cell(row=row_idx, column=id_write_col + 1).value
            if ext_id_val:
                data["externalId"] = str(ext_id_val).strip()

        expected_text = ""
        if exp_col is not None:
            exp_val = rd.cell(row=row_idx, column=exp_col + 1).value
            expected_text = str(exp_val or "")

        raw_actual = data["response_body"]

        verdict, reason = evaluate_simple(expected_text, raw_actual)

        verdicts[verdict] = verdicts.get(verdict, 0) + 1
        display_name = source_key if source_key != tc_name else tc_name
        print(f"  [{verdict:7s}] {display_name[:65]}")

        if dry_run:
            continue

        # Track per-row verdict for Google Sheets cell coloring after upload
        if sts_col is not None:
            row_verdicts[row_idx] = verdict

        # Write actual result: response JSON body
        if act_col is not None:
            cell = ws.cell(row=row_idx, column=act_col + 1)
            cell.value = raw_actual
            cell.alignment = _wrap_align()
            cell.font = _black_font()
            cell.fill = _white_fill()

        # Write verdict (PASS/FAIL/N/A) to "Kết quả hiện tại" (status) column
        if sts_col is not None:
            cell = ws.cell(row=row_idx, column=sts_col + 1)
            cell.value = verdict
            cell.alignment = _wrap_align()
            cell.font = _black_font()
            cell.fill = _white_fill()

        _autofit_row_height(ws, row_idx, act_col)

    return matched, not_found, verdicts, row_verdicts


def apply_cell_background_colors(spreadsheet_id: str, sheet_title: str,
                                  sts_col_1based: int, row_verdicts: dict,
                                  max_data_row: int, creds):
    """
    Apply background colors to status cells after upload to Google Sheets.

    PASS → #b6d7a8 (green), FAIL → #f4cccc (red), N/A → white.
    Uses Google Sheets API batchUpdate with repeatCell per cell for precision.
    """
    from googleapiclient.discovery import build

    sheets = build("sheets", "v4", credentials=creds, cache_discovery=False)

    # Resolve sheetId from title
    meta = sheets.spreadsheets().get(
        spreadsheetId=spreadsheet_id,
    ).execute()
    sheet_id = None
    for s in meta.get("sheets", []):
        props = s.get("properties", {})
        if props.get("title") == sheet_title:
            sheet_id = props.get("sheetId")
            break
    if sheet_id is None:
        print(f"  [color] Could not find sheet '{sheet_title}' — skipping background color")
        return

    color_map = {
        "PASS": {"red": 0.714, "green": 0.843, "blue": 0.659},
        "FAIL": {"red": 0.957, "green": 0.800, "blue": 0.800},
        "N/A":  {"red": 1.0,  "green": 1.0,  "blue": 1.0},
    }

    # row_verdicts: {sheet_idx_0based: {row_1based: verdict}}
    sheet_verdicts = row_verdicts.get(0, {})
    if not sheet_verdicts:
        print(f"  [color] No verdicts to color")
        return

    requests = []
    for row_1based, verdict in sheet_verdicts.items():
        if row_1based > max_data_row:
            continue
        color = color_map.get(verdict, color_map["N/A"])
        requests.append({
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": row_1based - 1,
                    "endRowIndex": row_1based,
                    "startColumnIndex": sts_col_1based - 1,
                    "endColumnIndex": sts_col_1based,
                },
                "cell": {
                    "userEnteredFormat": {
                        "backgroundColor": color,
                        "textFormat": {"bold": False},
                    }
                },
                "fields": "userEnteredFormat(backgroundColor,textFormat)"
            }
        })

    if requests:
        for i in range(0, len(requests), 100):
            chunk = requests[i:i+100]
            sheets.spreadsheets().batchUpdate(
                spreadsheetId=spreadsheet_id,
                body={"requests": chunk}
            ).execute()
        print(f"  [color] Applied background to {len(requests)} cells")


def merge_results(
    source_path: str,
    target_url: str,
    structure_path: str,
    dry_run: bool = False,
    auto_fix: bool = True,
    new_spreadsheet_name: str = None,
) -> dict:
    import openpyxl

    # --- Step 0: Google auth & download target ---
    spreadsheet_id = parse_spreadsheet_id(target_url)
    print(f"[0/5] Authenticating with Google ...")
    creds = get_google_credentials()

    tmpdir = tempfile.mkdtemp(prefix="merge_postman_")
    target_xlsx = os.path.join(tmpdir, "target.xlsx")

    print(f"[1/5] Downloading template Google Sheet: {spreadsheet_id}")
    download_gsheet_as_xlsx(spreadsheet_id, creds, target_xlsx)

    # --- Step 1: Load source ---
    print(f"[2/5] Loading source: {source_path}")
    source_results = load_source(source_path)
    print(f"      {len(source_results)} test results found")

    # --- Step 2: Load structure ---
    print(f"[3/5] Loading structure: {structure_path}")
    with open(structure_path, "r", encoding="utf-8") as f:
        structure = json.load(f)

    # --- Step 3: Open downloaded target xlsx ---
    print(f"[4/5] Loading target xlsx: {target_xlsx}")
    wb = openpyxl.load_workbook(target_xlsx)
    ws = wb.worksheets[0]
    # Also open a read-only copy with data_only=True to resolve formula values
    wb_ro = openpyxl.load_workbook(target_xlsx, data_only=True)
    ws_ro = wb_ro.worksheets[0]
    print(f"      Sheet: '{ws.title}'")

    col_map = detect_columns(ws_ro, structure)
    header_row    = structure.get("headerRow", 1)
    data_start_row = structure.get("dataStartRow", header_row + 1)

    tc_col  = col_map.get("testCaseName")
    exp_col = col_map.get("expectedResults")
    act_col = col_map.get("actualResult")
    sts_col = col_map.get("status")

    missing = [f for f, c in [("testCaseName", tc_col)] if c is None]
    if missing:
        raise ValueError(
            f"Cannot find columns: {missing}. "
            "Check structure.json columnMapping or target file header row."
        )

    print(f"      Column map: testCaseName={tc_col} status={sts_col} "
          f"expectedResults={exp_col} actualResult={act_col}")
    if sts_col is None:
        print("      [warn] 'status' column not found — Pass/Fail will not be written.")

    sample_names = []
    for r in ws.iter_rows(min_row=data_start_row, max_row=data_start_row + 3, values_only=True):
        if tc_col is not None and r[tc_col]:
            sample_names.append(str(r[tc_col])[:60])
    if sample_names:
        print(f"      Sample testCaseName values (rows {data_start_row}+): {sample_names}")

    # --- Step 4: Merge ---
    print(f"[5/5] Merging results (rows {data_start_row}+) ...")
    matched, not_found, verdicts, row_verdicts = _run_merge_loop(
        ws, source_results, col_map, data_start_row, dry_run, ws_readonly=ws_ro
    )

    print()
    print(f"  Matched : {matched} / {len(source_results)}")
    print(f"  PASS    : {verdicts.get('PASS', 0)}")
    print(f"  FAIL    : {verdicts.get('FAIL', 0)}")
    print(f"  N/A     : {verdicts.get('N/A', 0)}")
    if not_found:
        print(f"  Unmatched ({len(not_found)}): {not_found[:8]}"
              + (" ..." if len(not_found) > 8 else ""))

    # --- Auto-fix: if nothing matched, scan for the real header row ---
    if matched == 0 and auto_fix and len(source_results) > 0:
        print("\n[auto-fix] matched = 0 — scanning worksheet for correct header row ...")
        new_header, new_data_start, new_col_map = auto_detect_structure(ws_ro, structure)

        # Only adopt auto-detected values if they are actually better.
        # We check whether testCaseName and externalId resolved to non-empty values
        # in the new mapping (the original structure.json had ext_col=5 correct).
        resolved_tc   = new_col_map.get("testCaseName") is not None
        resolved_ext  = new_col_map.get("externalId") is not None
        any_better    = resolved_tc or resolved_ext

        if not any_better:
            print("[auto-fix] No better column mapping found — keeping original.")
            new_col_map = col_map.copy()
            new_header  = header_row
            new_data_start = data_start_row
        else:
            if new_header != header_row:
                print(f"[auto-fix] headerRow: {header_row} → {new_header}")
            if new_data_start != data_start_row:
                print(f"[auto-fix] dataStartRow: {data_start_row} → {new_data_start}")
            for field in ("testCaseName", "expectedResults", "actualResult", "status", "externalId"):
                old_c = col_map.get(field)
                new_c = new_col_map.get(field)
                if old_c != new_c:
                    print(f"[auto-fix] {field}: col {old_c} → {new_c}")

            header_row     = new_header
            data_start_row = new_data_start
            col_map        = new_col_map

        # Re-merge with corrected mapping
        print(f"\n[auto-fix] Re-merging (rows {data_start_row}+) ...")
        matched2, not_found2, verdicts2, row_verdicts2 = _run_merge_loop(
            ws, source_results, col_map, data_start_row, dry_run, ws_readonly=ws_ro
        )
        matched = matched2
        not_found = not_found2
        verdicts = verdicts2
        row_verdicts = row_verdicts2

    print()
    print(f"  Matched : {matched} / {len(source_results)}")
    print(f"  PASS    : {verdicts.get('PASS', 0)}")
    print(f"  FAIL    : {verdicts.get('FAIL', 0)}")
    print(f"  N/A     : {verdicts.get('N/A', 0)}")
    if not_found:
        print(f"  Unmatched ({len(not_found)}): {not_found[:8]}"
              + (" ..." if len(not_found) > 8 else ""))

    if not dry_run:
        # Update Evidence sheet
        print("\nUpdating Evidence sheet ...")
        update_evidence_sheet(wb, source_results)

        # Save merged xlsx locally
        merged_xlsx = os.path.join(tmpdir, "merged.xlsx")
        wb.save(merged_xlsx)
        print(f"Saved locally → {merged_xlsx}")

        # Verify before uploading
        print("\n[Verify] Re-reading saved file ...")
        verify = _verify_output(merged_xlsx, structure, col_map, data_start_row, matched)
        _print_verify(verify)

        # Create brand-new Google Sheet (target is kept untouched)
        new_name = new_spreadsheet_name or (
            f"Merged Results {time.strftime('%Y-%m-%d %H:%M:%S')}"
        )
        print(f"\n[Upload] Creating new Google Sheet: '{new_name}' ...")
        sheet_url, spreadsheet_id = create_new_gsheet_from_xlsx(merged_xlsx, new_name, creds)

        # Apply PASS/FAIL/N/A cell background colors after upload
        if sts_col is not None:
            sheet_title = ws.title
            max_data_row = ws.max_row
            apply_cell_background_colors(
                spreadsheet_id, sheet_title,
                sts_col + 1,  # convert 0-based to 1-based
                {0: row_verdicts},
                max_data_row, creds
            )

        print(f"\nDone! → {sheet_url}")
    else:
        print("\n[dry-run] No changes written.")
        sheet_url = None

    return {
        "matched": matched,
        "total_source": len(source_results),
        "verdicts": verdicts,
        "not_found": not_found,
        "spreadsheet_url": sheet_url,
    }


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Merge auto-postman results into a Google Spreadsheet"
    )
    parser.add_argument("--source",    required=True,
                        help="Path to api_results_*.xlsx (Postman output)")
    parser.add_argument("--target",    required=True,
                        help="Google Sheets URL or spreadsheet ID (used as template, not overwritten)")
    parser.add_argument("--structure", required=True,
                        help="Path to structure.json")
    parser.add_argument("--dry-run",   action="store_true",
                        help="Print what would happen without writing anything")
    parser.add_argument("--name",
                        help="Name of the new Google Sheet to create (default: auto-generated)")
    args = parser.parse_args()

    try:
        result = merge_results(
            source_path=args.source,
            target_url=args.target,
            structure_path=args.structure,
            dry_run=args.dry_run,
            new_spreadsheet_name=args.name,
        )
        print()
        print(json.dumps(result, ensure_ascii=False, indent=2))
    except Exception as exc:
        print(f"\n[error] {exc}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
