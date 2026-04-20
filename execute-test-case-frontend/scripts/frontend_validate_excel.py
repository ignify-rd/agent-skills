#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Prepare and finalize frontend validate test execution workbooks.

The script keeps spreadsheet and Excel work deterministic so Playwright MCP only
does browser work. It accepts either a Google Sheets link, a local .xlsx file, or
a JSON file containing a 2D values array.
"""

from __future__ import annotations

import argparse
import copy
import json
import os
import re
import shutil
import sys
import unicodedata
import urllib.parse
import urllib.request
from datetime import datetime
from pathlib import Path
from typing import Any

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
except ImportError as exc:  # pragma: no cover - dependency guard
    print(
        json.dumps(
            {
                "error": "Missing dependency",
                "detail": str(exc),
                "install": "pip install openpyxl Pillow",
            },
            ensure_ascii=False,
        ),
        file=sys.stderr,
    )
    sys.exit(2)


DEFAULT_OUT_DIR = "frontend_validate_runs"
DEFAULT_SHEET = "TestCase"

FIXED_COLUMNS = {
    "name": 2,
    "preconditions": 3,
    "importance": 4,
    "steps": 5,
    "data": 6,
    "expected": 7,
    "actual": 8,
    "result": 9,
}

LABEL_TO_KEY = {
    "name": "name",
    "test case name": "name",
    "testcase name": "name",
    "name testcase": "name",
    "test case title": "name",
    "summary": "name",
    "preconditions": "preconditions",
    "pre conditions": "preconditions",
    "pre-condition": "preconditions",
    "pre-conditions": "preconditions",
    "dieu kien bat buoc": "preconditions",
    "importance": "importance",
    "priority": "importance",
    "muc do uu tien": "importance",
    "step": "steps",
    "steps": "steps",
    "test step": "steps",
    "test steps": "steps",
    "cac buoc thuc hien": "steps",
    "data": "data",
    "data test": "data",
    "test data": "data",
    "du lieu test": "data",
    "expected result": "expected",
    "expected results": "expected",
    "ket qua mong muon": "expected",
    "actual result": "actual",
    "actual results": "actual",
    "ket qua thuc te": "actual",
    "result": "result",
    "test results": "result",
    "ket qua hien tai": "result",
    "lan 1": "result",
    "lan 1 chrome": "result",
    "chrome": "result",
    "status": "result",
}

GROUP_HEADER_LABELS = {
    "test suite",
    "test case",
    "attachments",
    "custom fields",
    "steps",
    "coverage",
    "time management",
    "result",
    "kich ban kiem thu",
    "ten man hinh ten chuc nang",
}

NON_VALIDATE_MAIN_SUITES = (
    "kiem tra chuc nang",
    "kiem tra logic",
    "kiem tra luong chinh",
    "kiem tra main flow",
    "kiem tra happy path",
    "kiem tra ngoai le",
    "kiem tra phan quyen",
    "kiem tra timeout",
    "kiem tra token",
    "kiem tra endpoint",
    "kiem tra search",
    "kiem tra tim kiem",
)

NON_VALIDATE_MAIN_TOKENS = (
    "chuc nang",
    "logic",
    "luong chinh",
    "main flow",
    "happy path",
    "ngoai le",
    "phan quyen",
    "timeout",
    "token",
    "endpoint",
)


def now_stamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def norm_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r", " ").replace("\n", " ").strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def display_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def is_validate_text(value: Any) -> bool:
    text = norm_text(value)
    return "kiem tra validate" in text or bool(re.search(r"\bvalidate\b", text))


def is_non_validate_main_suite(value: Any) -> bool:
    text = norm_text(value)
    if any(text.startswith(item) for item in NON_VALIDATE_MAIN_SUITES):
        return True
    is_kiem_tra = text.startswith("kiem tra") or text.startswith("ki m tra")
    is_field_suite = text.startswith("kiem tra truong") or text.startswith("ki m tra truong")
    if not is_kiem_tra or is_field_suite:
        return False
    tail = re.sub(r"^(kiem tra|ki m tra)\s+", "", text).strip()
    return any(tail == token or tail.startswith(f"{token} ") for token in NON_VALIDATE_MAIN_TOKENS)


def is_header_label(value: Any) -> bool:
    text = norm_text(value)
    return text in LABEL_TO_KEY or text in GROUP_HEADER_LABELS


def safe_filename(value: str, fallback: str = "case") -> str:
    text = unicodedata.normalize("NFKD", value or fallback)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"[^A-Za-z0-9._-]+", "_", text).strip("._")
    return (text or fallback)[:120]


def col_letter(col: int) -> str:
    result = ""
    while col:
        col, rem = divmod(col - 1, 26)
        result = chr(65 + rem) + result
    return result


def cell_value(row: list[Any], col_1based: int | None) -> Any:
    if not col_1based or col_1based < 1:
        return None
    idx = col_1based - 1
    return row[idx] if idx < len(row) else None


def parse_sheet_url(url: str) -> tuple[str | None, str | None]:
    match = re.search(r"/spreadsheets/d/([^/]+)", url)
    if not match:
        return None, None
    parsed = urllib.parse.urlparse(url)
    qs = urllib.parse.parse_qs(parsed.query)
    gid = qs.get("gid", [None])[0]
    return match.group(1), gid


def download_google_sheet(url: str, output_path: Path, gid: str | None = None) -> str:
    sheet_id, parsed_gid = parse_sheet_url(url)
    if not sheet_id:
        export_url = url
    else:
        gid = gid or parsed_gid
        query = {"format": "xlsx"}
        if gid:
            query["gid"] = gid
        export_url = (
            f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?"
            f"{urllib.parse.urlencode(query)}"
        )

    request = urllib.request.Request(
        export_url,
        headers={"User-Agent": "Mozilla/5.0 frontend-validate-excel/1.0"},
    )
    try:
        with urllib.request.urlopen(request, timeout=90) as response:
            data = response.read()
    except Exception as exc:
        raise RuntimeError(f"Cannot download Google Sheet export: {exc}") from exc

    if not data.startswith(b"PK"):
        preview = data[:300].decode("utf-8", errors="replace")
        raise RuntimeError(
            "Downloaded content is not an .xlsx file. The sheet may require "
            f"authentication or export permission. Preview: {preview}"
        )

    output_path.write_bytes(data)
    return export_url


def workbook_from_values_json(values_json: Path, output_path: Path, sheet_name: str) -> None:
    raw = json.loads(values_json.read_text(encoding="utf-8"))
    if isinstance(raw, dict):
        rows = raw.get("values") or raw.get("data") or raw.get("rows")
    else:
        rows = raw
    if not isinstance(rows, list):
        raise ValueError("values JSON must be a 2D array or an object with values/data/rows")

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name or DEFAULT_SHEET
    for r_idx, row in enumerate(rows, start=1):
        if not isinstance(row, list):
            continue
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx).value = value
    wb.save(output_path)


def key_for_label(value: Any) -> str | None:
    text = norm_text(value)
    if not text or text in GROUP_HEADER_LABELS:
        return None
    if text in LABEL_TO_KEY:
        return LABEL_TO_KEY[text]
    for label, key in LABEL_TO_KEY.items():
        if len(label) >= 8 and label in text:
            return key
    return None


def detect_columns(ws) -> tuple[int, dict[str, int]]:
    best_score = -1
    best_row = 1
    best_mapping: dict[str, int] = {}
    max_header_row = min(ws.max_row, 40)

    for row_idx in range(1, max_header_row + 1):
        mapping: dict[str, int] = {}
        for col_idx in range(1, ws.max_column + 1):
            key = key_for_label(ws.cell(row=row_idx, column=col_idx).value)
            if key and key not in mapping:
                mapping[key] = col_idx

        score = 0
        for key, weight in (
            ("name", 4),
            ("steps", 3),
            ("expected", 3),
            ("actual", 2),
            ("result", 2),
            ("preconditions", 1),
            ("data", 1),
        ):
            if key in mapping:
                score += weight
        if score > best_score:
            best_score = score
            best_row = row_idx
            best_mapping = mapping

    if best_score < 6:
        return 1, dict(FIXED_COLUMNS)

    columns = dict(FIXED_COLUMNS)
    columns.update(best_mapping)
    return best_row, columns


def copy_header_style(ws, src_col: int, dst_col: int, header_row: int) -> None:
    src = ws.cell(row=header_row, column=src_col)
    dst = ws.cell(row=header_row, column=dst_col)
    if src.has_style:
        dst._style = copy.copy(src._style)
    dst.font = copy.copy(src.font)
    dst.fill = copy.copy(src.fill)
    dst.border = copy.copy(src.border)
    dst.alignment = copy.copy(src.alignment)
    dst.number_format = src.number_format


def find_header_col(ws, header_row: int, normalized_labels: set[str]) -> int | None:
    for col in range(1, ws.max_column + 1):
        if norm_text(ws.cell(row=header_row, column=col).value) in normalized_labels:
            return col
    return None


def ensure_output_columns(ws, header_row: int, columns: dict[str, int]) -> dict[str, int]:
    output = dict(columns)
    if not output.get("actual"):
        output["actual"] = FIXED_COLUMNS["actual"]
    if not output.get("result"):
        output["result"] = FIXED_COLUMNS["result"]

    headers = {
        "actual": "Actual Result",
        "result": "Lần 1 (Chrome)",
    }
    for key, label in headers.items():
        cell = ws.cell(row=header_row, column=output[key])
        if not display_text(cell.value):
            cell.value = label

    existing_screenshot = find_header_col(ws, header_row, {"screenshot", "evidence"})
    existing_error = find_header_col(ws, header_row, {"error", "loi"})
    existing_executed = find_header_col(ws, header_row, {"executed at", "thoi gian chay"})

    next_col = max(ws.max_column + 1, output["result"] + 1)
    for key, existing, label in (
        ("screenshot", existing_screenshot, "Screenshot"),
        ("error", existing_error, "Error"),
        ("executedAt", existing_executed, "Executed At"),
    ):
        if existing:
            output[key] = existing
            continue
        output[key] = next_col
        ws.cell(row=header_row, column=next_col).value = label
        copy_header_style(ws, output["result"], next_col, header_row)
        next_col += 1

    return output


def is_pending(value: Any) -> bool:
    text = norm_text(value)
    return text in {
        "",
        "pending",
        "pendding",
        "todo",
        "not run",
        "not executed",
        "not execute",
        "chua chay",
        "chua thuc hien",
        "chua execute",
    }


def looks_like_suite_header(
    row: list[Any],
    name: str,
    steps: str,
    expected: str,
    data: str,
) -> bool:
    first = first_nonempty(row)
    if not first:
        return False
    if not name:
        return True
    if (not steps and not expected and not data) and (
        is_validate_text(name)
        or is_non_validate_main_suite(name)
        or norm_text(name).startswith("kiem tra truong")
        or ":" in name
    ):
        return True
    return False


def first_nonempty(row: list[Any]) -> str:
    for value in row:
        text = display_text(value)
        if text:
            return text
    return ""


def field_group_from(suite: str, name: str) -> str:
    suite_norm = norm_text(suite)
    if suite and not is_validate_text(suite) and not is_non_validate_main_suite(suite):
        return suite

    name_text = display_text(name)
    parts = [p for p in re.split(r"[_/|>-]+", name_text) if p.strip()]
    for idx, part in enumerate(parts):
        if is_validate_text(part) and idx + 1 < len(parts):
            return parts[idx + 1].strip()

    match = re.search(r"validate[_\s-]+([^_\-]+)", name_text, flags=re.IGNORECASE)
    if match:
        return match.group(1).strip()
    return suite or "Validate"


def collect_cases(
    ws,
    header_row: int,
    columns: dict[str, int],
    include_done: bool,
) -> list[dict[str, Any]]:
    cases: list[dict[str, Any]] = []
    current_suite = ""
    inside_validate = False
    max_col = max(ws.max_column, *(columns.values()))

    for row_idx in range(header_row + 1, ws.max_row + 1):
        row = [ws.cell(row=row_idx, column=col).value for col in range(1, max_col + 1)]
        if not any(display_text(v) for v in row):
            continue

        name = display_text(cell_value(row, columns.get("name")))
        steps = display_text(cell_value(row, columns.get("steps")))
        expected = display_text(cell_value(row, columns.get("expected")))
        data = display_text(cell_value(row, columns.get("data")))
        preconditions = display_text(cell_value(row, columns.get("preconditions")))
        importance = display_text(cell_value(row, columns.get("importance")))
        status = cell_value(row, columns.get("result"))

        if looks_like_suite_header(row, name, steps, expected, data):
            current_suite = first_nonempty(row)
            if is_validate_text(current_suite):
                inside_validate = True
            elif is_non_validate_main_suite(current_suite):
                inside_validate = False
            continue

        if not name or is_header_label(name):
            continue

        validate_case = inside_validate or is_validate_text(name) or is_validate_text(current_suite)
        if not validate_case:
            continue
        if not include_done and not is_pending(status):
            continue

        cases.append(
            {
                "id": f"r{row_idx}",
                "row": row_idx,
                "suite": current_suite,
                "fieldGroup": field_group_from(current_suite, name),
                "name": name,
                "preconditions": preconditions,
                "importance": importance,
                "steps": steps,
                "data": data,
                "expected": expected,
                "status": display_text(status),
                "safeName": safe_filename(f"{row_idx}_{name}", f"case_{row_idx}"),
                "actualCol": columns["actual"],
                "resultCol": columns["result"],
                "screenshotCol": columns["screenshot"],
                "errorCol": columns["error"],
                "executedAtCol": columns["executedAt"],
            }
        )

    return cases


def create_run(args: argparse.Namespace) -> dict[str, Any]:
    timestamp = now_stamp()
    out_dir = Path(args.out_dir or DEFAULT_OUT_DIR).resolve()
    run_dir = out_dir / f"frontend_validate_{timestamp}"
    run_dir.mkdir(parents=True, exist_ok=True)
    screenshots_dir = run_dir / "screenshots"
    screenshots_dir.mkdir(exist_ok=True)

    input_path = run_dir / "input.xlsx"
    source = ""
    export_url = ""
    if args.sheet_url:
        source = args.sheet_url
        export_url = download_google_sheet(args.sheet_url, input_path, gid=args.gid)
    elif args.xlsx:
        src = Path(args.xlsx).resolve()
        if not src.exists():
            raise FileNotFoundError(src)
        source = str(src)
        shutil.copy2(src, input_path)
    elif args.values_json:
        src = Path(args.values_json).resolve()
        if not src.exists():
            raise FileNotFoundError(src)
        source = str(src)
        workbook_from_values_json(src, input_path, args.sheet_name or DEFAULT_SHEET)
    else:
        raise ValueError("Provide --sheet-url, --xlsx, or --values-json")

    result_path = run_dir / f"frontend_validate_results_{timestamp}.xlsx"
    shutil.copy2(input_path, result_path)

    read_wb = openpyxl.load_workbook(input_path, data_only=True)
    sheet_name = args.sheet_name or (DEFAULT_SHEET if DEFAULT_SHEET in read_wb.sheetnames else read_wb.sheetnames[0])
    if sheet_name not in read_wb.sheetnames:
        raise ValueError(f"Sheet not found: {sheet_name}. Available: {read_wb.sheetnames}")
    read_ws = read_wb[sheet_name]
    header_row, detected_columns = detect_columns(read_ws)

    result_wb = openpyxl.load_workbook(result_path)
    result_ws = result_wb[sheet_name]
    output_columns = ensure_output_columns(result_ws, header_row, detected_columns)
    result_ws.freeze_panes = result_ws.freeze_panes or f"A{header_row + 1}"
    for key in ("actual", "screenshot", "error"):
        result_ws.column_dimensions[col_letter(output_columns[key])].width = max(
            result_ws.column_dimensions[col_letter(output_columns[key])].width or 0,
            30 if key != "actual" else 55,
        )
    result_wb.save(result_path)

    cases = collect_cases(read_ws, header_row, output_columns, include_done=args.include_done)

    cases_path = run_dir / "cases.json"
    cases_path.write_text(json.dumps(cases, ensure_ascii=False, indent=2), encoding="utf-8")

    run = {
        "timestamp": timestamp,
        "source": source,
        "exportUrl": export_url,
        "runDir": str(run_dir),
        "inputWorkbook": str(input_path),
        "resultWorkbook": str(result_path),
        "sheetName": sheet_name,
        "headerRow": header_row,
        "columns": output_columns,
        "casesFile": str(cases_path),
        "screenshotsDir": str(screenshots_dir),
        "caseCount": len(cases),
        "includeDone": bool(args.include_done),
    }
    run_path = run_dir / "run.json"
    run["runFile"] = str(run_path)
    run_path.write_text(json.dumps(run, ensure_ascii=False, indent=2), encoding="utf-8")
    return run


def normalize_verdict(value: Any, error: Any = None) -> str:
    text = norm_text(value)
    if text in {"p", "pass", "passed"}:
        return "P"
    if text in {"f", "fail", "failed"}:
        return "F"
    if text in {"e", "error", "err"}:
        return "E"
    return "E" if error else "F"


def load_results(args: argparse.Namespace) -> list[dict[str, Any]]:
    if args.results_file:
        raw = Path(args.results_file).read_text(encoding="utf-8")
    elif args.results_json:
        raw = args.results_json
    else:
        raw = sys.stdin.read()
    data = json.loads(raw)
    if isinstance(data, dict):
        data = data.get("results", data.get("data", []))
    if not isinstance(data, list):
        raise ValueError("Results must be a JSON array or an object with results/data")
    return data


def style_status_cell(cell, verdict: str) -> None:
    colors = {
        "P": ("E2F0D9", "008000"),
        "F": ("FCE4D6", "C00000"),
        "E": ("FFF2CC", "9C6500"),
    }
    fill, font = colors.get(verdict, ("FFFFFF", "000000"))
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = Font(color=font, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def ensure_summary_sheet(wb, run: dict[str, Any], counts: dict[str, int], result_count: int) -> None:
    if "Summary" in wb.sheetnames:
        del wb["Summary"]
    ws = wb.create_sheet("Summary", 0)
    rows = [
        ("Frontend Validate Test Summary", ""),
        ("Run timestamp", run["timestamp"]),
        ("Source", run.get("source", "")),
        ("Input workbook", run.get("inputWorkbook", "")),
        ("Result workbook", run.get("resultWorkbook", "")),
        ("Sheet", run.get("sheetName", "")),
        ("Validate cases selected", run.get("caseCount", 0)),
        ("Results received", result_count),
        ("PASS (P)", counts.get("P", 0)),
        ("FAIL (F)", counts.get("F", 0)),
        ("ERROR (E)", counts.get("E", 0)),
    ]
    for row_idx, (label, value) in enumerate(rows, start=1):
        ws.cell(row=row_idx, column=1).value = label
        ws.cell(row=row_idx, column=2).value = value
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].fill = PatternFill("solid", fgColor="4472C4")
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 100


def add_evidence_sheet(wb, evidence_rows: list[dict[str, Any]]) -> None:
    if "Evidence" in wb.sheetnames:
        del wb["Evidence"]
    ws = wb.create_sheet("Evidence")
    headers = ["#", "Source Row", "Field Group", "Test Case", "Verdict", "Actual", "Screenshot"]
    header_fill = PatternFill("solid", fgColor="4472C4")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 45
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 60
    ws.column_dimensions["G"].width = 80
    ws.freeze_panes = "A2"

    target_height = 260
    for idx, item in enumerate(evidence_rows, start=2):
        ws.cell(row=idx, column=1).value = idx - 1
        ws.cell(row=idx, column=2).value = item.get("row", "")
        ws.cell(row=idx, column=3).value = item.get("fieldGroup", "")
        ws.cell(row=idx, column=4).value = item.get("name", "")
        verdict_cell = ws.cell(row=idx, column=5)
        verdict_cell.value = item.get("verdict", "")
        style_status_cell(verdict_cell, item.get("verdict", ""))
        ws.cell(row=idx, column=6).value = item.get("actual", "")
        ws.cell(row=idx, column=6).alignment = Alignment(wrap_text=True, vertical="top")
        screenshot = item.get("screenshot", "")
        ws.cell(row=idx, column=7).value = screenshot
        ws.cell(row=idx, column=7).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[idx].height = target_height * 0.75 + 8
        if screenshot and os.path.exists(screenshot):
            try:
                img = XLImage(screenshot)
                ratio = target_height / img.height if img.height else 1
                img.height = target_height
                img.width = int(img.width * ratio)
                ws.add_image(img, f"G{idx}")
                ws.cell(row=idx, column=7).value = ""
            except Exception as exc:
                ws.cell(row=idx, column=7).value = f"{screenshot}\nEmbed error: {exc}"


def finalize_run(args: argparse.Namespace) -> dict[str, Any]:
    run_dir = Path(args.run_dir).resolve()
    run_path = run_dir / "run.json"
    if not run_path.exists():
        raise FileNotFoundError(run_path)
    run = json.loads(run_path.read_text(encoding="utf-8"))
    cases = json.loads(Path(run["casesFile"]).read_text(encoding="utf-8"))
    cases_by_id = {case["id"]: case for case in cases}
    cases_by_row = {int(case["row"]): case for case in cases}
    results = load_results(args)

    wb = openpyxl.load_workbook(run["resultWorkbook"])
    ws = wb[run["sheetName"]]
    cols = run["columns"]
    executed_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    counts = {"P": 0, "F": 0, "E": 0}
    evidence_rows: list[dict[str, Any]] = []

    for result in results:
        row_num = result.get("row")
        case = None
        if result.get("id") in cases_by_id:
            case = cases_by_id[result["id"]]
        elif row_num is not None:
            case = cases_by_row.get(int(row_num))
        if not case:
            continue

        row_num = int(case["row"])
        error = result.get("error", "")
        verdict = normalize_verdict(result.get("verdict", result.get("result")), error)
        actual = display_text(
            result.get("actual")
            or result.get("actualResult")
            or result.get("message")
            or result.get("observed")
            or error
        )
        screenshot = result.get("screenshot") or result.get("screenshotPath") or ""
        if isinstance(screenshot, list):
            screenshot = screenshot[0] if screenshot else ""
        screenshot = display_text(screenshot)

        ws.cell(row=row_num, column=cols["actual"]).value = actual
        ws.cell(row=row_num, column=cols["actual"]).alignment = Alignment(wrap_text=True, vertical="top")
        status_cell = ws.cell(row=row_num, column=cols["result"])
        status_cell.value = verdict
        style_status_cell(status_cell, verdict)
        ws.cell(row=row_num, column=cols["screenshot"]).value = screenshot
        if screenshot:
            ws.cell(row=row_num, column=cols["screenshot"]).hyperlink = screenshot
        ws.cell(row=row_num, column=cols["error"]).value = display_text(error)
        ws.cell(row=row_num, column=cols["executedAt"]).value = result.get("executedAt") or executed_at

        counts[verdict] = counts.get(verdict, 0) + 1
        evidence_rows.append(
            {
                "row": row_num,
                "fieldGroup": case.get("fieldGroup", ""),
                "name": case.get("name", ""),
                "verdict": verdict,
                "actual": actual,
                "screenshot": screenshot,
            }
        )

    ensure_summary_sheet(wb, run, counts, len(results))
    add_evidence_sheet(wb, evidence_rows)
    wb.save(run["resultWorkbook"])
    return {
        "resultWorkbook": run["resultWorkbook"],
        "runDir": str(run_dir),
        "totalSelected": run.get("caseCount", 0),
        "resultsReceived": len(results),
        "pass": counts.get("P", 0),
        "fail": counts.get("F", 0),
        "error": counts.get("E", 0),
    }


def print_run_summary(run: dict[str, Any]) -> None:
    cases = json.loads(Path(run["casesFile"]).read_text(encoding="utf-8"))
    groups: dict[str, int] = {}
    for case in cases:
        groups[case.get("fieldGroup") or "Validate"] = groups.get(case.get("fieldGroup") or "Validate", 0) + 1
    output = {
        "runDir": run["runDir"],
        "inputWorkbook": run["inputWorkbook"],
        "resultWorkbook": run["resultWorkbook"],
        "sheetName": run["sheetName"],
        "headerRow": run["headerRow"],
        "casesFile": run["casesFile"],
        "screenshotsDir": run["screenshotsDir"],
        "caseCount": run["caseCount"],
        "groups": groups,
    }
    print(json.dumps(output, ensure_ascii=False, indent=2))


def main() -> None:
    parser = argparse.ArgumentParser(description="Frontend validate Excel helper")
    sub = parser.add_subparsers(dest="command", required=True)

    p_prepare = sub.add_parser("prepare", help="Download/copy input and build run files")
    source = p_prepare.add_mutually_exclusive_group(required=True)
    source.add_argument("--sheet-url", help="Google Sheets URL to export as .xlsx")
    source.add_argument("--xlsx", help="Local .xlsx input file")
    source.add_argument("--values-json", help="2D values JSON exported from Sheets MCP")
    p_prepare.add_argument("--gid", help="Google Sheets gid override")
    p_prepare.add_argument("--sheet-name", help="Worksheet/tab name")
    p_prepare.add_argument("--out-dir", help=f"Output directory (default: {DEFAULT_OUT_DIR})")
    p_prepare.add_argument(
        "--include-done",
        action="store_true",
        help="Include cases that already have P/F/E or another result value",
    )

    p_finalize = sub.add_parser("finalize", help="Write MCP execution results to workbook")
    p_finalize.add_argument("--run-dir", required=True, help="Run directory from prepare")
    result_source = p_finalize.add_mutually_exclusive_group()
    result_source.add_argument("--results-file", help="JSON file returned by browser_run_code")
    result_source.add_argument("--results-json", help="Raw JSON returned by browser_run_code")

    args = parser.parse_args()
    try:
        if args.command == "prepare":
            run = create_run(args)
            print_run_summary(run)
        elif args.command == "finalize":
            print(json.dumps(finalize_run(args), ensure_ascii=False, indent=2))
    except Exception as exc:
        print(json.dumps({"error": type(exc).__name__, "detail": str(exc)}, ensure_ascii=False), file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
